import os
import os
import tempfile
import unittest
from datetime import datetime
from unittest.mock import patch
from pathlib import Path

from ComplaintSearch.claims_excel import ExcelClaimsSearcher
from openpyxl import Workbook, load_workbook


class ExcelClaimsSearchTest(unittest.TestCase):
    """Tests for ExcelClaimsSearcher.search."""

    def _create_file(self, path: str, headers=None) -> None:
        wb = Workbook()
        ws = wb.active
        if headers is None:
            headers = ["complaint", "customer", "subject", "part_code", "date"]
        ws.append(headers)
        ws.append(["noise", "ACME", "engine", "X1", datetime(2023, 1, 1)])
        ws.append(["crack", "BETA", "body", "X2", datetime(2022, 5, 1)])
        wb.save(path)

    def test_search_filters(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            searcher = ExcelClaimsSearcher(file_path)
            result = searcher.search({"customer": "ACME"}, year=2023)
            self.assertEqual(len(result), 1)
            self.assertEqual(result[0]["customer"], "ACME")
            empty = searcher.search({"customer": "ACME"}, year=2022)
            self.assertEqual(empty, [])

    def test_normalization_and_fuzzy(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)

            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(["\u015fikayet var", "GAMMA", "door", "X3", datetime(2023, 2, 1)])
            wb.save(file_path)

            searcher = ExcelClaimsSearcher(file_path)

            accent = searcher.search({"complaint": "sikayet"})
            self.assertEqual(len(accent), 1)
            self.assertEqual(accent[0]["customer"], "GAMMA")

            typo = searcher.search({"complaint": "noize"})
            self.assertEqual(len(typo), 1)
            self.assertEqual(typo[0]["customer"], "ACME")

    def test_turkish_headers(self) -> None:
        """Ensure search works when headers are Turkish."""
        headers = [
            "müşteri şikayeti",
            "müşteri",
            "konu",
            "parça kodu",
            "tarih",
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path, headers=headers)
            searcher = ExcelClaimsSearcher(file_path)
            result = searcher.search({"müşteri": "ACME"}, year=2023)
            self.assertEqual(len(result), 1)
            self.assertIn("musteri", result[0])
            self.assertEqual(result[0]["musteri"], "ACME")

    def test_hata_tarihi_header(self) -> None:
        """Year filters should work with 'Hata Tarihi' header."""
        headers = [
            "müşteri şikayeti",
            "müşteri",
            "konu",
            "parça kodu",
            "hata tarihi",
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path, headers=headers)
            searcher = ExcelClaimsSearcher(file_path)
            result = searcher.search({"müşteri": "ACME"}, year=2023)
            self.assertEqual(len(result), 1)
            self.assertEqual(result[0]["musteri"], "ACME")
            empty = searcher.search({"müşteri": "ACME"}, year=2022)
            self.assertEqual(empty, [])

    def test_dynamic_header_detection(self) -> None:
        """Headers should be detected after initial metadata rows."""
        file_path = Path("CC/F160_Customer_Claims.xlsx")
        searcher = ExcelClaimsSearcher(file_path)
        result = searcher.search({"PPM Adet": 1})
        self.assertTrue(len(result) > 0)
        for rec in result:
            self.assertEqual(rec["ppm adet"], 1)

    def test_unique_values(self) -> None:
        """``unique_values`` should return sorted distinct entries."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)

            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(["extra", "ACME", "engine", "X1", datetime(2024, 1, 1)])
            wb.save(file_path)

            searcher = ExcelClaimsSearcher(file_path)
            customers = searcher.unique_values("customer")
            self.assertEqual(customers, ["ACME", "BETA"])

    def test_env_path_used(self) -> None:
        """File path should come from ``COMPLAINTS_XLSX_PATH`` when not provided."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            with patch.dict("os.environ", {"COMPLAINTS_XLSX_PATH": file_path}):
                searcher = ExcelClaimsSearcher()
                result = searcher.search({"customer": "ACME"}, year=2023)
                self.assertEqual(len(result), 1)

    def test_missing_file_raises(self) -> None:
        """A missing Excel file should raise ``FileNotFoundError``."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "missing.xlsx")
            searcher = ExcelClaimsSearcher(file_path)
            with self.assertRaises(FileNotFoundError):
                searcher.search({})

    def test_year_range(self) -> None:
        """Records should be filterable by a start and end year."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "claims.xlsx")
            self._create_file(file_path)
            searcher = ExcelClaimsSearcher(file_path)
            both = searcher.search({}, start_year=2022, end_year=2023)
            self.assertEqual(len(both), 2)
            just_2022 = searcher.search({}, start_year=2022, end_year=2022)
            self.assertEqual(len(just_2022), 1)
            overlap = searcher.search({}, year=2023, start_year=2022, end_year=2023)
            self.assertEqual(len(overlap), 1)


if __name__ == "__main__":
    unittest.main()
